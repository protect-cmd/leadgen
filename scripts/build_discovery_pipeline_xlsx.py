"""Build the Tenant Discovery Pipeline Excel workbook.

Generates docs/Tenant_Discovery_Pipeline.xlsx with:
- Discovery Pipeline tab pre-populated with each builder's starting queue
- Legend tab with column rules
- Dropdown validation on Stage, Classification, and Y/N/TBD columns
- Conditional formatting (stuck rows red, lead's queue yellow, Live green, archived gray)
- Frozen header row + protected Lead notes column
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = Path(__file__).resolve().parent.parent / "docs" / "Tenant_Discovery_Pipeline.xlsx"

HEADERS = [
    "Builder",
    "State",
    "County / Source",
    "Source URL",
    "Stage",
    "Classification",
    "Date-enumerable?",
    "Tenant name exposed?",
    "Property/defendant address exposed?",
    "Est. weekly filings",
    "Evidence",
    "Blocker",
    "Last updated",
    "Lead notes",
    "PR link",
]

ROWS = [
    # Nourul / Texas
    ["Nourul", "TX", "Travis County JP Court (Austin)", "https://odysseypa.traviscountytx.gov/JPPublicAccess/", "Researching", "TBD", "TBD", "TBD", "TBD", "", "", "Tyler Odyssey on .gov — Bright Data blocks. Needs Playwright with residential IP. Confirm Forcible Detainer hearing type label first.", "2026-05-27", "", ""],
    ["Nourul", "TX", "Williamson County JP Court", "https://judicialrecords.wilco.org", "Researching", "TBD", "TBD", "TBD", "TBD", "", "", "Same Tyler stack as Travis. Test public vs securepa subdomain for JP records.", "2026-05-27", "", ""],
    ["Nourul", "TX", "Fort Bend County JP Court", "https://tylerpaw.fortbendcountytx.gov/PublicAccess/", "Researching", "TBD", "TBD", "TBD", "TBD", "", "", "Tyler PAW. Civil/eviction Date Filed behavior unconfirmed. Portal intermittently 522.", "2026-05-27", "", ""],
    ["Nourul", "TX", "Montgomery County JP Courts", "", "Researching", "TBD", "TBD", "TBD", "TBD", "", "", "No public extract found in prior sweep. Probe for date-enumerable JP search.", "2026-05-27", "", ""],
    ["Nourul", "TX", "Brazoria County JP Courts", "", "Researching", "TBD", "TBD", "TBD", "TBD", "", "", "Unmapped. Probe for public extract.", "2026-05-27", "", ""],
    ["Nourul", "TX", "Collin County JP Courts", "", "Researching", "TBD", "TBD", "TBD", "TBD", "", "", "Unmapped. Probe for public extract.", "2026-05-27", "", ""],
    ["Nourul", "TX", "Denton County JP Courts", "", "Researching", "TBD", "TBD", "TBD", "TBD", "", "", "Unmapped. Probe for public extract.", "2026-05-27", "", ""],
    ["Nourul", "TX", "Galveston County JP Courts", "", "Researching", "TBD", "TBD", "TBD", "TBD", "", "", "Unmapped. Probe for public extract.", "2026-05-27", "", ""],
    ["Nourul", "TX", "El Paso County JP Courts", "", "Researching", "TBD", "TBD", "TBD", "TBD", "", "", "Unmapped. Probe for public extract.", "2026-05-27", "", ""],
    # Franz / Ohio
    ["Franz", "OH", "Cuyahoga County (Cleveland Housing Court)", "https://clevelandhousingcourt.org/accessible-civil-docket", "Researching", "Yellow", "Y", "Y", "N", 290, "", "No property address exposed — yellow until SearchBug enrichment proven on Ohio addresses.", "2026-05-27", "Static HTML docket ~583 cases/2-weeks confirmed 2026-05-14.", ""],
    ["Franz", "OH", "Montgomery County (Dayton Municipal Court)", "https://clerkofcourt.daytonohio.gov", "Researching", "Yellow", "Y", "TBD", "TBD", "", "", "Filing Date is explicit search type. Address exposure unconfirmed. Use Dayton clerk portal (avoid PRO system).", "2026-05-27", "", ""],
    ["Franz", "OH", "Summit County (Akron Municipal Court)", "", "Researching", "TBD", "TBD", "TBD", "TBD", "", "", "Online docket exists but no bulk/CSV export found in prior sweep. Re-probe.", "2026-05-27", "", ""],
    ["Franz", "OH", "Lucas County (Toledo Municipal Court)", "", "Researching", "TBD", "TBD", "TBD", "TBD", "", "", "Online docket exists but no bulk/CSV export found. Re-probe.", "2026-05-27", "", ""],
    ["Franz", "OH", "Butler County Municipal Court", "", "Researching", "TBD", "TBD", "TBD", "TBD", "", "", "Unmapped in detail. Probe online docket.", "2026-05-27", "", ""],
    ["Franz", "OH", "Stark County Municipal Court", "", "Researching", "TBD", "TBD", "TBD", "TBD", "", "", "Unmapped in detail. Probe online docket.", "2026-05-27", "", ""],
    ["Franz", "OH", "Lorain County Municipal Court", "", "Researching", "TBD", "TBD", "TBD", "TBD", "", "", "Unmapped in detail. Probe online docket.", "2026-05-27", "", ""],
    # Donnald / Tennessee
    ["Donnald", "TN", "Shelby County General Sessions (Memphis)", "https://shelbygeneralsessions.com/115/Download-Case-Information", "Researching", "Yellow", "TBD", "TBD", "TBD", "", "", "HIGHEST UPSIDE. Anomalous Download Case Info page — could be free bulk extract or clerk-gated. Test in browser before any build.", "2026-05-27", "", ""],
    ["Donnald", "TN", "Knox County Civil Sessions (Knoxville)", "https://knoxcounty.org/civil/dockets.php", "Researching", "Yellow", "Y", "Y", "N", "", "", "Weekly PDF dockets no login. No address in PDFs. Paid sub for case detail. Yellow until enrichment.", "2026-05-27", "", ""],
    ["Donnald", "TN", "Hamilton County General Sessions (Chattanooga)", "https://edockets.us/hamiltontn/", "Researching", "Yellow", "Y", "Y", "N", "", "", "Clean JSON API for docket PDFs by date. Detainer type confirmed. No address in any docket format.", "2026-05-27", "", ""],
    ["Donnald", "TN", "Sumner County General Sessions (Gallatin)", "https://sumnercourts.com/", "Researching", "Yellow", "Y", "Y", "N", "", "", "Direct daily PDF dockets by courtroom. No login. No address in any field.", "2026-05-27", "", ""],
    ["Donnald", "TN", "Rutherford County General Sessions (Murfreesboro)", "", "Researching", "TBD", "TBD", "TBD", "TBD", "", "", "Unmapped. Probe for public docket or download.", "2026-05-27", "", ""],
    ["Donnald", "TN", "Williamson County General Sessions (Franklin)", "", "Researching", "TBD", "TBD", "TBD", "TBD", "", "", "Unmapped. Probe for public docket or download.", "2026-05-27", "", ""],
    ["Donnald", "TN", "Montgomery County General Sessions (Clarksville)", "", "Researching", "TBD", "TBD", "TBD", "TBD", "", "", "Unmapped. Probe for public docket or download.", "2026-05-27", "", ""],
]

LEGEND_ROWS = [
    ("Builder", "Nourul / Franz / Donnald"),
    ("State", "Two-letter state code (TX / OH / TN / AZ / CA / KS / IN)"),
    ("County / Source", "County name + court level (e.g. Travis County JP Court)"),
    ("Source URL", "Live link to the portal — builder fills as soon as they find it"),
    ("Stage", "Researching → Classified-pending-approval → Approved-to-build → Building → Submitted-for-review → Live (terminal: Rejected, Skipped, Upgrade-proposed)"),
    ("Classification", "Green / Yellow / Red / TBD. Green requires tenant name + property/defendant address."),
    ("Date-enumerable?", "Y / N / TBD — can you list filings by filing date with no party name, no CAPTCHA, no login?"),
    ("Tenant name exposed?", "Y / N / TBD"),
    ("Property/defendant address exposed?", "Y / N / TBD — required for Green; Yellow if only tenant name"),
    ("Est. weekly filings", "Number from live probe (count rows over 7-day lookback). Leave blank if unmeasured."),
    ("Evidence", "Link to screenshot / probe script / sample CSV / Loom — anything lead can click to verify"),
    ("Blocker", "Free text. Empty = no blocker. Update when stuck."),
    ("Last updated", "YYYY-MM-DD. Builders touch this daily even if nothing changed. Stale > 2 days = stuck signal."),
    ("Lead notes", "Lead-only column. Builders do not edit."),
    ("PR link", "GitHub pull request URL — filled when stage moves to Submitted-for-review"),
]

STAGE_VALUES = "Researching,Classified-pending-approval,Approved-to-build,Building,Submitted-for-review,Live,Rejected,Skipped,Upgrade-proposed"
CLASSIFICATION_VALUES = "Green,Yellow,Red,TBD"
YN_VALUES = "Y,N,TBD"

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
LEAD_NOTES_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def build_workbook() -> Workbook:
    wb = Workbook()
    pipe = wb.active
    pipe.title = "Discovery Pipeline"

    # Header row
    pipe.append(HEADERS)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = pipe.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    pipe.row_dimensions[1].height = 32

    # Data rows
    for row in ROWS:
        pipe.append(row)

    # Column widths
    widths = {
        "A": 11, "B": 7, "C": 42, "D": 50, "E": 26, "F": 14,
        "G": 17, "H": 19, "I": 32, "J": 18, "K": 28, "L": 60,
        "M": 14, "N": 40, "O": 30,
    }
    for col, width in widths.items():
        pipe.column_dimensions[col].width = width

    # Freeze header
    pipe.freeze_panes = "A2"

    # Data validations
    n_rows = len(ROWS) + 200  # leave headroom for new rows
    stage_dv = DataValidation(type="list", formula1=f'"{STAGE_VALUES}"', allow_blank=True)
    cls_dv = DataValidation(type="list", formula1=f'"{CLASSIFICATION_VALUES}"', allow_blank=True)
    yn_dv = DataValidation(type="list", formula1=f'"{YN_VALUES}"', allow_blank=True)

    pipe.add_data_validation(stage_dv)
    pipe.add_data_validation(cls_dv)
    pipe.add_data_validation(yn_dv)

    stage_dv.add(f"E2:E{n_rows}")
    cls_dv.add(f"F2:F{n_rows}")
    yn_dv.add(f"G2:I{n_rows}")

    # Conditional formatting on data range A2:O{n_rows}
    rng = f"A2:O{n_rows}"

    stuck_rule = FormulaRule(
        formula=[f'AND($M2<>"", TODAY()-$M2>2, NOT(OR($E2="Live", $E2="Rejected", $E2="Skipped")))'],
        fill=PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
    )
    queue_rule = FormulaRule(
        formula=[f'OR($E2="Classified-pending-approval", $E2="Submitted-for-review")'],
        fill=PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
    )
    live_rule = FormulaRule(
        formula=[f'$E2="Live"'],
        fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    )
    archived_rule = FormulaRule(
        formula=[f'AND($F2="Red", $E2="Skipped")'],
        fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    )

    pipe.conditional_formatting.add(rng, stuck_rule)
    pipe.conditional_formatting.add(rng, queue_rule)
    pipe.conditional_formatting.add(rng, live_rule)
    pipe.conditional_formatting.add(rng, archived_rule)

    # Tint Lead notes column on data rows for visual cue
    for row_idx in range(2, len(ROWS) + 2):
        pipe.cell(row=row_idx, column=14).fill = LEAD_NOTES_FILL

    # Wrap text on long columns
    for row_idx in range(2, len(ROWS) + 2):
        for col_letter in ("C", "D", "L", "N"):
            pipe[f"{col_letter}{row_idx}"].alignment = Alignment(wrap_text=True, vertical="top")

    # Enable autofilter
    pipe.auto_filter.ref = f"A1:O{len(ROWS) + 1}"

    # Legend tab
    legend = wb.create_sheet("Legend")
    legend.append(["Column", "Allowed values / rule"])
    for cell in legend[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in LEGEND_ROWS:
        legend.append(row)
    legend.column_dimensions["A"].width = 36
    legend.column_dimensions["B"].width = 90
    for row_idx in range(2, len(LEGEND_ROWS) + 2):
        legend[f"B{row_idx}"].alignment = Alignment(wrap_text=True, vertical="top")
        legend[f"A{row_idx}"].font = Font(bold=True)
    legend.freeze_panes = "A2"

    return wb


def main() -> None:
    wb = build_workbook()
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
